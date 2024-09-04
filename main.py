
#### PAGOS DE CONSTRUCCION ####

import base_datos_conexion as db
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from datetime import datetime, date, timedelta
from openpyxl import Workbook
from typing import List
from tkcalendar import Calendar
import os
import sys

class Asistencia():
    def __init__(self,estado,fecha):
        self.estado = estado
        self.fecha = fecha

    def imprimir(self):
        print("los datos de la asistencia son: ")
        print(self.fecha)
        print(self.estado)

class ResultadoExcel():
    def __init__(self,num_semana,lista_dias,total):
        self.num_semana = num_semana
        self.lista_dias: List[Asistencia] = lista_dias
        self.total = total

    def obtener_lista_numero_semana(self):
        lista_dias_ = []

        for dia in self.lista_dias:
            dia_semana_num = dia.fecha.weekday()
            if dia.estado == 'si_trabajo':
                lista_dias_.append(dia_semana_num)

        return lista_dias_

class DiaNumeroSemana():
    def __init__(self, num_semana, fecha, estado):
        self.num_semana = num_semana
        self.fecha = fecha
        self.estado = estado

class SemanaPagada():
    def __init__(self, id, numero_semana, fecha_registro_pago, estado_pago):
        self.id = id
        self.numero_semana = numero_semana
        self.fecha_registro_pago = fecha_registro_pago
        self.estado_pago = estado_pago


def obtener_ruta_absoluta(relativa):
    """ Obtiene la ruta absoluta para archivos empaquetados con PyInstaller. """
    try:
        base_path = sys._MEIPASS  # PyInstaller guarda los recursos aquí al empaquetar
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relativa)

def main():
    mensaje_label_dias = None
    mensaje_label_pag = None
    #def registrar_pago_si():
        #fecha_actual = datetime.now().date()
        #insetar_asistencia(fecha_actual,'si_trabajo')


    # Calcula pagos pendientes y los inserta en la base de datos
    def pago_num_semana_estado():
        lista_semanas = []  
        semana_pagos = []
        
        fecha_minima = db.buscar_fecha_minima()
        fecha_maxima = db.buscar_fecha_maxima()

        if fecha_minima[0][0] and fecha_maxima[0][0]:
            num_semana_minima = obtener_numero_semana(fecha_minima[0][0])
            num_semana_maxima = obtener_numero_semana(fecha_maxima[0][0])

            datos_semanas = db.obtener_semanas_pagadas()
            print("222222222233333333333333333")
            print(datos_semanas)
            print("222222222233333333333333333")
            for datos in datos_semanas:
                numero_semana = datos[1]
                lista_semanas.append(numero_semana)
        
            for i in range(num_semana_minima, num_semana_maxima + 1):
                if i in lista_semanas:
                    continue
                semana_pago = SemanaPagada(None, i, None, 'no_pagado')

                db.insertar_pago(semana_pago)

    # Datos para mostrar en la interfaz
    def obtener_pagos_pendientes():
        lista_valores_pag = []
        datos_semanas = db.obtener_semanas_pagadas()
        fecha_actual = datetime.now().date()
        num_semana_actual = obtener_numero_semana(fecha_actual)

        for datos in datos_semanas:
            id_semanas_pagadas = datos[0]
            numero_semana = datos[1]
            estado_pago = datos[3]
            if estado_pago != 'si_pagado' and numero_semana < num_semana_actual:
                lista_valores_pag.append(SemanaPagada(id_semanas_pagadas, numero_semana, None, estado_pago))
            
        print("00000000000000000000000")
        print(lista_valores_pag)
        print("00000000000000000000000")
        return lista_valores_pag

    def obtener_fecha_inicio_final_numero_semana(year, week_number):
        
        first_day_of_year = datetime(year, 1, 1)
        
        # Averiguar el día de la semana del primer día del año (lunes=0, domingo=6)
        first_day_of_year_weekday = first_day_of_year.weekday()

        # Calcular el primer lunes del año (la primera semana puede no comenzar en lunes)
        days_to_first_monday = (7 - first_day_of_year_weekday) % 7
        first_monday = first_day_of_year + timedelta(days=days_to_first_monday)

        # Calcular la fecha de inicio de la semana
        start_of_week = first_monday + timedelta(weeks=week_number - 1)

        # Calcular la fecha de fin de la semana
        end_of_week = start_of_week + timedelta(days=6)

        return start_of_week.date(), end_of_week.date()  

    def registrar_i():
        pass

    def insetar_asistencia(fecha, estado):
        dia_semana = fecha.weekday()

        if dia_semana == 5 or dia_semana == 6: # 5 Sabado, 6 Domingo
            print("No se puede registrar asistencia de Sabados y Domingos")
            return

        nueva_asistencia = Asistencia(estado, fecha)
        fecha_coincidencias = db.buscar_asistencia_por_fecha(fecha)
        
        if not fecha_coincidencias:
            db.insertar(nueva_asistencia)
            nueva_asistencia.imprimir()
            mensaje_label.config(text="Asistencia Registrada!!!!") #agregado
        else:
            print("Ya hay un registro")
            mensaje_label.config(text="Ya hay un registro!!!!!")#agregado


    def registrar_si():
        fecha_actual = datetime.now().date()
        insetar_asistencia(fecha_actual,'si_trabajo')
        cont_no_paga, total_pagar = dias_trabajados_total_pago()
        mensaje_label_dias.config(text=str(cont_no_paga))
        mensaje_label_pag.config(text=str(total_pagar))


    def registrar_no():
        fecha_actual = datetime.now().date()
        insetar_asistencia(fecha_actual,'no_trabajo')

    def obtener_numero_semana(fecha):
        fecha_str = str(fecha)
        fechas_dt = datetime.strptime(fecha_str, '%Y-%m-%d')
        semana = fechas_dt.isocalendar()
        numero_semana = semana[1]
        return numero_semana

    def generar_reporte():
        lista_tmp = []
        lista_resultado_excel = []
        lista_asistencias = []
        datos = db.obtener_todas_asistencias()


        for dato in datos:
            lista_asistencias.append(Asistencia(dato[2], dato[1]))
        print("77777777777777778888888")
        print(lista_asistencias)
        print(type(lista_asistencias))
        print("7777777777777777888888")
        for registro in lista_asistencias:
            fecha_registro = registro.fecha
            estado_ = registro.estado
            print("7777777777777777")
            print(fecha_registro)
            print(type(fecha_registro))
            print("7777777777777777")
            num_semana = obtener_numero_semana(fecha_registro)
            lista_tmp.append(DiaNumeroSemana(num_semana, fecha_registro,estado_))

        for num_sem in range(1,53):
            dias_semana = []
            dias_trabajados = 0

            for registro in lista_tmp:
                if registro.num_semana == num_sem:
                    if registro.estado == 'si_trabajo':
                        dias_trabajados += 1
                    dias_semana.append(Asistencia(registro.estado, registro.fecha))

            if len(dias_semana) > 0:
                suma_semana = dias_trabajados * 80
                lista_resultado_excel.append(ResultadoExcel(num_sem,dias_semana,suma_semana))
            
        exportar_excel(lista_resultado_excel)
        print("0000000000000000000")
        print(lista_resultado_excel)
        print("0000000000000000000")

    def exportar_excel(datos):
        cont = 2
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "SEMANA"
        ws['B1'] = "LUNES"
        ws['C1'] = "MARTES"
        ws['D1'] = "MIERCOLES"
        ws['E1'] = "JUEVES"
        ws['F1'] = "VIERNES"
        ws['G1'] = "TOTAL"
        
        for registro in datos:
            ws[f'A{cont}'] = registro.num_semana
            
            lista_dias_ = registro.obtener_lista_numero_semana()

            for i in range(0,6): # 0 es lunes, 1 es martes ...
                letra = chr(ord('A')+i+1)
                if i in lista_dias_:
                    ws[f'{letra}{cont}'] = 'SI'
                else:
                    ws[f'{letra}{cont}'] = 'NO'
            ws[f'G{cont}'] = registro.total

            cont += 1


        wb.save("dias_semana.xlsx")
        print("Archivo 'dias_semana.xlsx' generado exitosamente.")

    def calcular_fechas_pendientes():
        fechas_registradas = set()
        fechas_pendientes = []
        asis_todos_dias = db.obtener_todas_asistencias()

        for dato in asis_todos_dias:
            fechas_registradas.add(dato[1])
        
        fecha_minima = db.buscar_fecha_minima()

        fecha_minima_date = fecha_minima[0][0]
        fecha_maxima_date  = datetime.now().date()
        fecha_nueva_date = fecha_minima_date

        if fecha_minima_date and fecha_maxima_date:
            fecha_nueva_date += timedelta(days=1)
            while fecha_nueva_date < fecha_maxima_date:
                dia_num = fecha_nueva_date.weekday()
                
                if dia_num == 5 or dia_num == 6:
                    fecha_nueva_date += timedelta(days=1)
                    continue
                if fecha_nueva_date not in fechas_registradas:
                    fechas_pendientes.append(str(fecha_nueva_date))
                fecha_nueva_date += timedelta(days=1)

        return fechas_pendientes 

    def dias_trabajados_total_pago():
        fechas_registradas = set()
        dias_trabajados = []
        cont_no_paga = 0
        semanas_no_pagadas_list = []
        asis_todos_dias = db.obtener_todas_asistencias()
        for dato in asis_todos_dias:
            if dato[2] == "si_trabajo":
                fechas_registradas.add(dato[1])

        semanas_no_pagadas_list = db.obtener_semanas_pagadas(flag_pagado=False)
        for dato in semanas_no_pagadas_list:
            for dias in fechas_registradas:
                numero_sem = obtener_numero_semana(dias)
                if dato[1] == numero_sem:
                    cont_no_paga += 1
        
        total_pagar = cont_no_paga * 80

        return cont_no_paga, total_pagar


    ventana = tk.Tk()
    ventana.title("Formulario de Registro")
    ventana.geometry("1000x1000")
    ventana.configure(bg="lightblue")

    tk.Label(ventana, text="Se trabajo hoy?", bg="#2755a2", fg="white").grid(row=1, column=0, padx=10, pady=10, sticky="nsew")

    boton_registrar_si = tk.Button(ventana, text="Si",bg="#2755a2", fg="white", command=registrar_si)
    boton_registrar_si.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")

    mensaje_label = tk.Label(ventana, text="") #agregado
    mensaje_label.grid(row=1, column=3, padx=10, pady=10, sticky="nsew") #agregado

    boton_registrar_no = tk.Button(ventana, text="No",bg="#2755a2", fg="white", command=registrar_no)
    boton_registrar_no.grid(row=1, column=2, padx=10, pady=10, sticky="nsew")

    boton_generar_reporte = tk.Button(ventana, text="Generar Reporte", command=generar_reporte, bg="#2755a2", fg="white")
    boton_generar_reporte.grid(row=2, column=3, padx=10, pady=10, sticky="nsew")


    tk.Label(ventana, text="Asistencias pendientes: ", bg="#2755a2", fg="white").grid(row=4, column=0, padx=10, pady=10, sticky="nsew")
    frame = tk.Frame(ventana)
    frame.place(relx=0.5, rely=0.5, anchor="center")
    frame.grid(row=5, column=0)

    # Crear un Canvas para simular la tabla y permitir el desplazamiento
    canvas = tk.Canvas(frame)
    canvas.pack(side="left", fill="both", expand=True)

    # Crear un Scrollbar vertical para el Canvas
    scrollbar = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    # Configurar el Canvas para usar el Scrollbar
    canvas.configure(yscrollcommand=scrollbar.set)
    canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    # Crear un frame secundario dentro del Canvas, con borde y fondo claro
    frame_in_canvas = tk.Frame(canvas, bd=2, relief="solid", bg="#e2e2e1")  # Añadir fondo claro
    canvas.create_window((0, 0), window=frame_in_canvas, anchor="nw")




    #calcular_fechas_pendientes()
    pago_num_semana_estado()
    def llenar_tabla(frame_in_canvas):
        # Primero, eliminar todas las filas existentes en la tabla
        for widget in frame_in_canvas.winfo_children():
            widget.destroy()

        valores = calcular_fechas_pendientes()

        label_fecha = tk.Label(frame_in_canvas, text="Fecha", bg="#e2e2e1", width=10)
        label_fecha.grid(row=0, column=0)

        for i, valor in enumerate(valores):
            label = tk.Label(frame_in_canvas, text=valor, width=15, anchor="w", bg="#e2e2e1")
            label.grid(row=i+1, column=0, padx=5, pady=2)
            boton_si = tk.Button(frame_in_canvas, text="SI",bg="#2755a2", fg="white", command=lambda i=i: si_button_click(valores[i]))
            boton_si.grid(row=i+1,columnspan=2, column=1, padx=5, pady=2)

            boton_no = tk.Button(frame_in_canvas, bg="#2755a2", fg="white", text="NO", command=lambda i=i: no_button_click(valores[i]))
            boton_no.grid(row=i+1,columnspan=2, column=4, padx=5, pady=2)
        frame_in_canvas.update_idletasks()

        
    # Función para el evento del botón
    def si_button_click(fecha_str):
        fecha_date = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        insetar_asistencia(fecha_date,'si_trabajo')
        cont_no_paga, total_pagar = dias_trabajados_total_pago()
        mensaje_label_dias.config(text=str(cont_no_paga))
        mensaje_label_pag.config(text=str(total_pagar))
        llenar_tabla(frame_in_canvas)

    def no_button_click(fecha_str):
        fecha_date = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        insetar_asistencia(fecha_date,'no_trabajo')
        llenar_tabla(frame_in_canvas)




    ###################################### Segunda Tabla
    tk.Label(ventana, text="Pagos pendientes: ", bg="#2755a2", fg="white").grid(row=6, column=0, padx=10, pady=10, sticky="nsew")
    frame_1 = tk.Frame(ventana)
    frame_1.place(relx=0.5, rely=0.5, anchor="center")
    frame_1.grid(row=7, column=0)

    # Crear un Canvas para simular la tabla y permitir el desplazamiento
    canvas_1 = tk.Canvas(frame_1)
    canvas_1.pack(side="left", fill="both", expand=True)

    # Crear un Scrollbar vertical para el Canvas
    scrollbar_1 = ttk.Scrollbar(frame_1, orient="vertical", command=canvas_1.yview)
    scrollbar_1.pack(side="right", fill="y")

    # Configurar el Canvas para usar el Scrollbar
    canvas_1.configure(yscrollcommand=scrollbar_1.set)
    canvas_1.bind('<Configure>', lambda e: canvas_1.configure(scrollregion=canvas_1.bbox("all")))

    # Crear un frame secundario dentro del Canvas, con borde y fondo claro
    frame_in_canvas_1 = tk.Frame(canvas_1, bd=2, relief="solid", bg="#e2e2e1")  # Añadir fondo claro
    canvas_1.create_window((0, 0), window=frame_in_canvas_1, anchor="nw")



    # Funcion para marcar un pago
    def si_button_pago(objeto):
        objeto.fecha_registro_pago = datetime.now().date()
        num_semana = objeto.numero_semana
        fechas_pendientes = calcular_fechas_pendientes()
        for fecha_str in fechas_pendientes:
            fecha_date = datetime.strptime(fecha_str, "%Y-%m-%d").date()
            numero_sem_asist_pendientes = obtener_numero_semana(fecha_date)
            if num_semana == numero_sem_asist_pendientes:
                mensaje_text = "No se puede realizar la operacion. Hay Asistencias pendientes por marcar!"
                mostrar_mensaje_validacion(mensaje_text)
                return # me saca inmediatamente de la funcion
            
        db.actualizar_estado_pagado(objeto)
        cont_no_paga, total_pagar = dias_trabajados_total_pago()
        mensaje_label_dias.config(text=str(cont_no_paga))
        mensaje_label_pag.config(text=str(total_pagar))
        llenar_tabla_pago(frame_in_canvas_1)

    ### Funcion para mostrar una nueva ventana emergente
    def mostrar_mensaje_validacion(mensaje_text):
        # Crear una nueva ventana
        ventana_validacion = tk.Toplevel(ventana)
        ventana_validacion.title("menssaje")
        ventana_validacion.geometry("500x150")
        
        # Crear un mensaje en la nueva ventana
        mensaje = tk.Label(ventana_validacion, text = mensaje_text, pady=20)
        mensaje.pack()

        # Botón para cerrar la ventana
        boton_cerrar = tk.Button(ventana_validacion, text="Cerrar", command=ventana_validacion.destroy)
        boton_cerrar.pack(pady=10)

        
    #calcular_fechas_pendientes()
    #pago_num_semana_estado()
    def llenar_tabla_pago(frame_in_canvas_2):
        # Primero, eliminar todas las filas existentes en la tabla
        for widget in frame_in_canvas_2.winfo_children():
            widget.destroy()

        lista_valores = obtener_pagos_pendientes()

        label_cab_pagos = tk.Label(frame_in_canvas_2, text="N de Semana", bg="#e2e2e1", width=15)
        label_cab_pagos.grid(row=0, column=0)

        label_cab_fecha = tk.Label(frame_in_canvas_2, text="Rango de Semana", bg="#e2e2e1", width=15)
        label_cab_fecha.grid(row=0, column=1)

        for i, valor in enumerate(lista_valores):
            label = tk.Label(frame_in_canvas_2, text=valor.numero_semana, width=15, anchor="w", bg="#e2e2e1")
            label.grid(row=i+1, column=0, padx=5, pady=2)

            fecha_inicial, fecha_final = obtener_fecha_inicio_final_numero_semana(2024, valor.numero_semana)
            fecha_junta = str(fecha_inicial) + " / " + str(fecha_final)

            label_2 = tk.Label(frame_in_canvas_2, text=fecha_junta, width=25, anchor="w", bg="#e2e2e1")
            label_2.grid(row=i+1, column=1, padx=5, pady=2)
            
            boton_si = tk.Button(frame_in_canvas_2, text="SI",bg="#2755a2", fg="white", command=lambda i=i: si_button_pago(lista_valores[i]))
            boton_si.grid(row=i+1, column=5, padx=5, pady=2)

        frame_in_canvas_2.update_idletasks()
    ######################################
    cont_no_paga, total_pagar = dias_trabajados_total_pago()

    tk.Label(ventana, text="Dias pendientes no pagados: ", bg="#2755a2", fg="white").grid(row=10, column=0, padx=10, pady=10, sticky="nsew")
    mensaje_label_dias = tk.Label(ventana, text=str(cont_no_paga), bg="#2755a2", fg="white") #agregado
    mensaje_label_dias.grid(row=10, column=1, padx=10, pady=10, sticky="nsew") #agregado

    tk.Label(ventana, text="Total pendiente a pagar: ", bg="#2755a2", fg="white").grid(row=11, column=0, padx=10, pady=10, sticky="nsew")
    mensaje_label_pag = tk.Label(ventana, text=str(total_pagar), bg="#2755a2", fg="white") #agregado
    mensaje_label_pag .grid(row=11, column=1, padx=10, pady=10, sticky="nsew") #agregado

    def mostrar_calendario():
        # Crear una nueva ventana para el calendario
        ventana_calendario = tk.Toplevel(ventana)
        ventana_calendario.title("Selecciona una fecha")

        # Crear un widget de calendario
        calendario = Calendar(ventana_calendario, selectmode="day")
        calendario.pack(pady=20)

        def seleccionar_fecha():
            # Obtener la fecha seleccionada y mostrarla en el campo de entrada
            fecha_seleccionada = calendario.get_date()
            fecha = datetime.strptime(fecha_seleccionada, "%m/%d/%y").date()
            entry_fecha.delete(0, tk.END)
            entry_fecha.insert(0, fecha)
            ventana_calendario.destroy()
            
        # Botón para seleccionar la fecha
        boton_seleccionar = ttk.Button(ventana_calendario, text="Seleccionar", command=seleccionar_fecha)
        boton_seleccionar.pack(pady=10)



    entry_fecha = ttk.Entry(ventana, width=20)
    entry_fecha.grid(row=4, column=3, padx=10, pady=20)

    def consultar_asistencia():
        fecha_calendar = entry_fecha.get()
        estado_trabajo = "No hay registros!!!!"
        if fecha_calendar:
            fecha = datetime.strptime(fecha_calendar, "%Y-%m-%d").date()
            fecha_bd = db.buscar_asistencia_por_fecha(fecha)
            for estado in fecha_bd:
                estado_trabajo = estado[2]
        
        if estado_trabajo:
            label_result.config(text=estado_trabajo)
        
        print(estado_trabajo)


    # Botón con un ícono de calendario
    ruta_imagen = obtener_ruta_absoluta(os.path.join('recursos', 'calendario_2.png'))
    icono_calendario = tk.PhotoImage(file=ruta_imagen)  # Reemplaza con la ruta de tu ícono
    boton_calendario = ttk.Button(ventana, image=icono_calendario, command=mostrar_calendario)
    boton_calendario.grid(row=4, column=4)

    boton_consultar = tk.Button(ventana, text="Buscar", command=consultar_asistencia, bg="#2755a2", fg="white")
    boton_consultar.grid(row=4, column=5, padx=10, pady=2, sticky="nsew")


    label_result = tk.Label(ventana, text="", anchor="w", bg="#e2e2e1")
    label_result.grid(row=5, column=5, padx=5, pady=2)

    llenar_tabla(frame_in_canvas)
    llenar_tabla_pago(frame_in_canvas_1)

    ventana.mainloop()

if __name__ == "__main__":
    main()