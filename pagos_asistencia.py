
#### PAGOS DE CONSTRUCCION ####
from base_datos_conexion import *
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from datetime import datetime, date, timedelta
from openpyxl import Workbook
from typing import List

class Asistencia():
    def __init__(self,estado,fecha):
        self.estado = estado
        self.fecha = fecha

    def imprimir(self):
        print("los datos de la asistencia son: ")
        print(self.fecha)
        print(self.estado)

class ResultadoExcel():
    def __init__(self,num_semana,lista_dias,total):
        self.num_semana = num_semana
        self.lista_dias: List[Asistencia] = lista_dias
        self.total = total

    def obtener_lista_numero_semana(self):
        lista_dias_ = []

        for dia in self.lista_dias:
            dia_semana_num = dia.fecha.weekday()
            if dia.estado == 'si_trabajo':
                lista_dias_.append(dia_semana_num)

        return lista_dias_

class DiaNumeroSemana():
    def __init__(self, num_semana, fecha, estado):
        self.num_semana = num_semana
        self.fecha = fecha
        self.estado = estado

class SemanaPagada():
    def __init__(self, id, numero_semana, fecha_registro_pago, estado_pago):
        self.id = id
        self.numero_semana = numero_semana
        self.fecha_registro_pago = fecha_registro_pago
        self.estado_pago = estado_pago

#def registrar_pago_si():
    #fecha_actual = datetime.now().date()
    #insetar_asistencia(fecha_actual,'si_trabajo')


# Calcula pagos pendientes y los inserta en la base de datos
def pago_num_semana_estado():
    lista_semanas = []  
    semana_pagos = []
    
    fecha_minima = buscar_fecha_minima()
    fecha_maxima = buscar_fecha_maxima()

    num_semana_minima = obtener_numero_semana(fecha_minima[0][0])
    num_semana_maxima = obtener_numero_semana(fecha_maxima[0][0])

    datos_semanas = obtener_semanas_pagadas()
    for datos in datos_semanas:
        numero_semana = datos[1]
        lista_semanas.append(numero_semana)
 
    for i in range(num_semana_minima, num_semana_maxima + 1):
        if i in lista_semanas:
            continue
        semana_pago = SemanaPagada(None, i, None, 'no_pagado')

        insertar_pago(semana_pago)

# Datos para mostrar en la interfaz
def obtener_pagos_pendientes():
    lista_valores_pag = []
    datos_semanas = obtener_semanas_pagadas()
    for datos in datos_semanas:
        id_semanas_pagadas = datos[0]
        numero_semana = datos[1]
        estado_pago = datos[3]
        if estado_pago != 'si_pagado':
            lista_valores_pag.append(SemanaPagada(id_semanas_pagadas, numero_semana, None, estado_pago))


    print("00000000000000000000000")
    print(lista_valores_pag)
    print("00000000000000000000000")
    return lista_valores_pag

def obtener_fecha_inicio_final_numero_semana(year, week_number):
    
    first_day_of_year = datetime(year, 1, 1)
    
    # Averiguar el día de la semana del primer día del año (lunes=0, domingo=6)
    first_day_of_year_weekday = first_day_of_year.weekday()

    # Calcular el primer lunes del año (la primera semana puede no comenzar en lunes)
    days_to_first_monday = (7 - first_day_of_year_weekday) % 7
    first_monday = first_day_of_year + timedelta(days=days_to_first_monday)

    # Calcular la fecha de inicio de la semana
    start_of_week = first_monday + timedelta(weeks=week_number - 1)

    # Calcular la fecha de fin de la semana
    end_of_week = start_of_week + timedelta(days=6)

    return start_of_week.date(), end_of_week.date()  

def registrar_i():
    pass

def insetar_asistencia(fecha, estado):
    dia_semana = fecha.weekday()

    if dia_semana == 5 or dia_semana == 6:
        print("No se puede registrar asistencia")
        return

    nueva_asistencia = Asistencia(estado, fecha)
    fecha_coincidencias = buscar_asistencia_por_fecha(fecha)
    
    if not fecha_coincidencias:
        insertar(nueva_asistencia)
        nueva_asistencia.imprimir()
    else:
        print("Ya hay un registro")




def registrar_si():
    fecha_actual = datetime.now().date()
    insetar_asistencia(fecha_actual,'si_trabajo')
    

def registrar_no():
    fecha_actual = datetime.now().date()
    insetar_asistencia(fecha_actual,'no_trabajo')

def obtener_numero_semana(fecha):
    fecha_str = str(fecha)
    fechas_dt = datetime.strptime(fecha_str, '%Y-%m-%d')
    semana = fechas_dt.isocalendar()
    numero_semana = semana[1]
    return numero_semana

def generar_reporte():
    lista_tmp = []
    lista_resultado_excel = []
    lista_asistencias = []
    datos = obtener_todas_asistencias()


    for dato in datos:
        lista_asistencias.append(Asistencia(dato[2], dato[1]))

    for registro in lista_asistencias:
        fecha_registro = registro.fecha
        estado_ = registro.estado
        num_semana = obtener_numero_semana(fecha_registro)
        lista_tmp.append(DiaNumeroSemana(num_semana, fecha_registro,estado_))

    for num_sem in range(1,36):
        dias_semana = []
        dias_trabajados = 0

        for registro in lista_tmp:
            if registro.num_semana == num_sem:
                if registro.estado == 'si_trabajo':
                    dias_trabajados += 1
                dias_semana.append(Asistencia(registro.estado, registro.fecha))

        if len(dias_semana) > 0:
            suma_semana = dias_trabajados * 80
            lista_resultado_excel.append(ResultadoExcel(num_sem,dias_semana,suma_semana))
        
    exportar_excel(lista_resultado_excel)
    print("0000000000000000000")
    print(lista_resultado_excel)
    print("0000000000000000000")

def exportar_excel(datos):
    cont = 2
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "SEMANA"
    ws['B1'] = "LUNES"
    ws['C1'] = "MARTES"
    ws['D1'] = "MIERCOLES"
    ws['E1'] = "JUEVES"
    ws['F1'] = "VIERNES"
    ws['G1'] = "TOTAL"
    
    for registro in datos:
        ws[f'A{cont}'] = registro.num_semana
        
        lista_dias_ = registro.obtener_lista_numero_semana()

        for i in range(0,6): # 0 es lunes, 1 es martes ...
            letra = chr(ord('A')+i+1)
            if i in lista_dias_:
                ws[f'{letra}{cont}'] = 'SI'
            else:
                ws[f'{letra}{cont}'] = 'NO'
        ws[f'G{cont}'] = registro.total

        cont += 1


    wb.save("dias_semana.xlsx")
    print("Archivo 'dias_semana.xlsx' generado exitosamente.")

def calcular_fechas_pendientes():
    fechas_registradas = set()
    fechas_pendientes = []
    asis_todos_dias = obtener_todas_asistencias()

    for dato in asis_todos_dias:
        fechas_registradas.add(dato[1])
    
    fecha_minima = buscar_fecha_minima()
    fecha_maxima = buscar_fecha_maxima()

    fecha_minima_date = fecha_minima[0][0]
    fecha_maxima_date = fecha_maxima[0][0]
    fecha_nueva_date = fecha_minima_date
    
    while fecha_nueva_date < fecha_maxima_date:
        fecha_nueva_date += timedelta(days=1)
        dia_num = fecha_nueva_date.weekday()
        if dia_num == 5 or dia_num == 6:
            continue
        if fecha_nueva_date not in fechas_registradas:
            fechas_pendientes.append(str(fecha_nueva_date))

    return fechas_pendientes 

ventana = tk.Tk()
ventana.title("Formulario de Registro")
ventana.geometry("1000x1000")


tk.Label(ventana, text="Se trabajo hoy?").grid(row=1, column=0, padx=10, pady=10, sticky="nsew")

boton_registrar_si = tk.Button(ventana, text="Si", command=registrar_si)
boton_registrar_si.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")

boton_registrar_no = tk.Button(ventana, text="No", command=registrar_no)
boton_registrar_no.grid(row=1, column=2, padx=10, pady=10, sticky="nsew")

boton_generar_reporte = tk.Button(ventana, text="Generar Reporte", command=generar_reporte)
boton_generar_reporte.grid(row=2, column=3, padx=10, pady=10, sticky="nsew")


tk.Label(ventana, text="Asistencias pendientes: ").grid(row=4, column=0, padx=10, pady=10, sticky="nsew")
frame = tk.Frame(ventana)
frame.place(relx=0.5, rely=0.5, anchor="center")
frame.grid(row=5, column=0)

# Crear un Canvas para simular la tabla y permitir el desplazamiento
canvas = tk.Canvas(frame)
canvas.pack(side="left", fill="both", expand=True)

# Crear un Scrollbar vertical para el Canvas
scrollbar = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
scrollbar.pack(side="right", fill="y")

# Configurar el Canvas para usar el Scrollbar
canvas.configure(yscrollcommand=scrollbar.set)
canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

# Crear un frame secundario dentro del Canvas, con borde y fondo claro
frame_in_canvas = tk.Frame(canvas, bd=2, relief="solid", bg="#0000FF")  # Añadir fondo claro
canvas.create_window((0, 0), window=frame_in_canvas, anchor="nw")




#calcular_fechas_pendientes()
#pago_num_semana_estado()
def llenar_tabla(frame_in_canvas):
    # Primero, eliminar todas las filas existentes en la tabla
    for widget in frame_in_canvas.winfo_children():
        widget.destroy()

    valores = calcular_fechas_pendientes()
    

    for i, valor in enumerate(valores):
        label = tk.Label(frame_in_canvas, text=valor, width=15, anchor="w", bg="#f0f0f0")
        label.grid(row=i, column=0, padx=5, pady=2)
        boton_si = tk.Button(frame_in_canvas, text="SI", command=lambda i=i: si_button_click(valores[i]))
        boton_si.grid(row=i,columnspan=2, column=1, padx=5, pady=2)

        boton_no = tk.Button(frame_in_canvas, text="NO", command=lambda i=i: no_button_click(valores[i]))
        boton_no.grid(row=i,columnspan=2, column=4, padx=5, pady=2)
    frame_in_canvas.update_idletasks()

    
# Función para el evento del botón
def si_button_click(fecha_str):
    fecha_date = datetime.strptime(fecha_str, "%Y-%m-%d").date()
    insetar_asistencia(fecha_date,'si_trabajo')
    llenar_tabla(frame_in_canvas)

def no_button_click(fecha_str):
    fecha_date = datetime.strptime(fecha_str, "%Y-%m-%d").date()
    insetar_asistencia(fecha_date,'no_trabajo')
    llenar_tabla(frame_in_canvas)




###################################### Segunda Tabla
tk.Label(ventana, text="Pagos pendientes: ").grid(row=6, column=0, padx=10, pady=10, sticky="nsew")
frame_1 = tk.Frame(ventana)
frame_1.place(relx=0.5, rely=0.5, anchor="center")
frame_1.grid(row=7, column=0)

# Crear un Canvas para simular la tabla y permitir el desplazamiento
canvas_1 = tk.Canvas(frame_1)
canvas_1.pack(side="left", fill="both", expand=True)

# Crear un Scrollbar vertical para el Canvas
scrollbar_1 = ttk.Scrollbar(frame_1, orient="vertical", command=canvas_1.yview)
scrollbar_1.pack(side="right", fill="y")

# Configurar el Canvas para usar el Scrollbar
canvas_1.configure(yscrollcommand=scrollbar_1.set)
canvas_1.bind('<Configure>', lambda e: canvas_1.configure(scrollregion=canvas_1.bbox("all")))

# Crear un frame secundario dentro del Canvas, con borde y fondo claro
frame_in_canvas_1 = tk.Frame(canvas_1, bd=2, relief="solid", bg="#0000FF")  # Añadir fondo claro
canvas_1.create_window((0, 0), window=frame_in_canvas_1, anchor="nw")



# Funcion para marcar un pago
def si_button_pago(objeto):
    objeto.fecha_registro_pago = datetime.now().date()
    actualizar_estado_pagado(objeto)
    llenar_tabla_pago(frame_in_canvas_1)



#calcular_fechas_pendientes()
#pago_num_semana_estado()
def llenar_tabla_pago(frame_in_canvas_2):
    # Primero, eliminar todas las filas existentes en la tabla
    for widget in frame_in_canvas_2.winfo_children():
        widget.destroy()

    lista_valores = obtener_pagos_pendientes()

    for i, valor in enumerate(lista_valores):
        label = tk.Label(frame_in_canvas_2, text=valor.numero_semana, width=15, anchor="w", bg="#f0f0f0")
        label.grid(row=i, column=0, padx=5, pady=2)

        fecha_inicial, fecha_final = obtener_fecha_inicio_final_numero_semana(2024, valor.numero_semana)
        fecha_junta = str(fecha_inicial) + " / " + str(fecha_final)

        label_2 = tk.Label(frame_in_canvas_2, text=fecha_junta, width=25, anchor="w", bg="#f0f0f0")
        label_2.grid(row=i, column=1, padx=5, pady=2)
        
        boton_si = tk.Button(frame_in_canvas_2, text="SI", command=lambda i=i: si_button_pago(lista_valores[i]))
        boton_si.grid(row=i, column=5, padx=5, pady=2)

    frame_in_canvas_2.update_idletasks()
######################################


llenar_tabla(frame_in_canvas)
llenar_tabla_pago(frame_in_canvas_1)

ventana.mainloop()



